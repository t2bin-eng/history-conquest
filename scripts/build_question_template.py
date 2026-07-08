from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = "templates/문제_업로드_양식.xlsx"

HEADER_FILL = PatternFill("solid", start_color="1D4ED8", end_color="1D4ED8")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BASE_FONT = Font(name="Arial")
EXAMPLE_FILL = PatternFill("solid", start_color="FEF9C3", end_color="FEF9C3")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# --- 안내 시트 ---
guide = wb.active
guide.title = "안내"
guide.sheet_view.showGridLines = False
guide.column_dimensions["A"].width = 90

lines = [
    ("역사 정복 — 문제 업로드 양식 안내", True, 14),
    ("", False, 11),
    ("1. '문제_양식' 시트에 문제를 입력하세요. 노란색 예시 행은 참고 후 삭제하거나 덮어써도 됩니다.", False, 11),
    ("2. 열 설명", True, 12),
    ("   - 카테고리: 대단원/시대 구분 (예: 선사·고대, 삼국시대, 고려, 조선, 근현대)", False, 11),
    ("   - 난이도: 하 / 중 / 고 중 하나 (드롭다운으로 선택)", False, 11),
    ("     · 하 = 변경(외곽) 지역용 쉬운 문제, 중 = 거점 지역용, 고 = 요충지용 어려운 문제", False, 11),
    ("   - 문제: 문제 본문 텍스트 (자료 지문이 있으면 문제 앞에 함께 적어주세요)", False, 11),
    ("   - 선택지1~5: 객관식 보기 5개 (반드시 5개 모두 채워주세요)", False, 11),
    ("   - 정답번호: 정답인 선택지 번호 1~5 중 하나 (드롭다운으로 선택)", False, 11),
    ("   - 제한시간(초): 비워두면 난이도별 기본값 적용(하 15초 / 중 18초 / 고 20초)", False, 11),
    ("", False, 11),
    ("3. 다 채운 뒤 교사 대시보드의 '문제 은행 관리'에서 이 파일을 업로드하면 됩니다.", False, 11),
    ("4. 정답 텍스트를 따로 옮겨 적지 않아도 되도록 '정답번호'만 선택하게 했습니다 — 오탈자로 정답이 어긋나는 것을 방지합니다.", False, 11),
]

row = 1
for text, bold, size in lines:
    cell = guide.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", bold=bold, size=size)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    row += 1

# --- 문제_양식 시트 ---
sheet = wb.create_sheet("문제_양식")
headers = [
    "카테고리", "난이도(하/중/고)", "문제",
    "선택지1", "선택지2", "선택지3", "선택지4", "선택지5",
    "정답번호(1~5)", "제한시간(초, 선택)",
]
widths = [14, 16, 45, 20, 20, 20, 20, 20, 14, 16]

for col, (title, width) in enumerate(zip(headers, widths), start=1):
    cell = sheet.cell(row=1, column=col, value=title)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    sheet.column_dimensions[cell.column_letter].width = width

sheet.freeze_panes = "A2"
sheet.row_dimensions[1].height = 28

example_rows = [
    ["고려", "고", "고려 시대에 제작된 세계 최초의 금속활자본은?", "직지심체요절", "팔만대장경", "삼국유사", "동의보감", "조선왕조실록", 1, ""],
    ["근현대", "중", "1920년 김좌진 장군이 일본군을 크게 물리친 전투는?", "청산리 전투", "봉오동 전투", "홍경성 전투", "간도 참변", "훈춘 사건", 1, ""],
]

for r, row_data in enumerate(example_rows, start=2):
    for c, value in enumerate(row_data, start=1):
        cell = sheet.cell(row=r, column=c, value=value)
        cell.font = BASE_FONT
        cell.fill = EXAMPLE_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))

# 빈 입력 행 (예시 2행 + 빈 행 148행 = 총 150행 분량 서식)
TOTAL_ROWS = 150
for r in range(len(example_rows) + 2, TOTAL_ROWS + 2):
    for c in range(1, len(headers) + 1):
        cell = sheet.cell(row=r, column=c)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))

last_row = TOTAL_ROWS + 1

difficulty_dv = DataValidation(type="list", formula1='"하,중,고"', allow_blank=True, showDropDown=False)
difficulty_dv.error = "하, 중, 고 중 하나를 선택하세요."
difficulty_dv.errorTitle = "잘못된 난이도"
sheet.add_data_validation(difficulty_dv)
difficulty_dv.add(f"B2:B{last_row}")

answer_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True, showDropDown=False)
answer_dv.error = "1, 2, 3, 4, 5 중 하나를 선택하세요."
answer_dv.errorTitle = "잘못된 정답번호"
sheet.add_data_validation(answer_dv)
answer_dv.add(f"I2:I{last_row}")

wb.save(OUT_PATH)
print(f"saved: {OUT_PATH}")
