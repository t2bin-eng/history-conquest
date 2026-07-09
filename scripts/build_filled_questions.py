from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

from questions_data import QUESTIONS

OUT_PATH = "templates/문제_100개_업로드용.xlsx"

DIFF_LABEL = {"LOW": "하", "MID": "중", "HIGH": "고"}

HEADER_FILL = PatternFill("solid", start_color="1D4ED8", end_color="1D4ED8")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BASE_FONT = Font(name="Arial")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()
sheet = wb.active
sheet.title = "문제_양식"

headers = [
    "카테고리", "난이도(하/중/고)", "문제",
    "선택지1", "선택지2", "선택지3", "선택지4", "선택지5",
    "정답번호(1~5)", "제한시간(초, 선택)",
]
widths = [14, 16, 55, 22, 22, 22, 22, 22, 14, 16]

for col, (title, width) in enumerate(zip(headers, widths), start=1):
    cell = sheet.cell(row=1, column=col, value=title)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    sheet.column_dimensions[cell.column_letter].width = width

sheet.freeze_panes = "A2"
sheet.row_dimensions[1].height = 28

questions_sorted = sorted(QUESTIONS, key=lambda q: q["num"])

for r, q in enumerate(questions_sorted, start=2):
    row_data = [
        q["category"],
        DIFF_LABEL[q["difficulty"]],
        q["text"],
        *q["choices"],
        q["answer"],
        "",
    ]
    for c, value in enumerate(row_data, start=1):
        cell = sheet.cell(row=r, column=c, value=value)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 3))

last_row = len(questions_sorted) + 1

difficulty_dv = DataValidation(type="list", formula1='"하,중,고"', allow_blank=True, showDropDown=False)
sheet.add_data_validation(difficulty_dv)
difficulty_dv.add(f"B2:B{last_row}")

answer_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True, showDropDown=False)
sheet.add_data_validation(answer_dv)
answer_dv.add(f"I2:I{last_row}")

wb.save(OUT_PATH)
print(f"saved: {OUT_PATH} ({len(questions_sorted)} rows)")
