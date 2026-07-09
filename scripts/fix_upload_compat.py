import sys
from openpyxl import Workbook, load_workbook

src_path = sys.argv[1]
dst_path = sys.argv[2]

src_wb = load_workbook(src_path)
src_sheet = src_wb[src_wb.sheetnames[0]]

dst_wb = Workbook()
dst_sheet = dst_wb.active
dst_sheet.title = "문제_양식"

for row in src_sheet.iter_rows():
    for cell in row:
        if cell.value is not None:
            dst_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

dst_wb.save(dst_path)
print(f"saved: {dst_path} (rows={src_sheet.max_row})")
